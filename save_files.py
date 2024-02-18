import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def make_xlsx(act, act_del, save_path):
    today = datetime.date.today()
    today_str = today.strftime("%d_%m_%y")
    act.to_excel(f'{save_path}/act_ref_{today_str}.xlsx', index=False)
    act_del.to_excel(f'{save_path}/act_del_{today_str}.xlsx', index=False)

    workbook = load_workbook(f'{save_path}/act_ref_{today_str}.xlsx')
    workbook = coloring_date(workbook, today)

    workbook.save(f'{save_path}/act_ref_{today_str}.xlsx')


def coloring_date(workbook, today):
    sheet = workbook['Sheet1']

    fill = PatternFill(start_color='FFFF00',
                       end_color='FFFF00',
                       fill_type='solid')

    for row in sheet.iter_rows(min_row=2):
        date_beer_value = row[-1].value
        delta = datetime.timedelta(180)
        if date_beer_value is not None:
            date_beer = datetime.datetime.strptime(str(date_beer_value),
                                                   "%d.%m.%Y").date()
            offset = today - date_beer
            if offset > delta:
                for cell in row:
                    cell.fill = fill

    return workbook
